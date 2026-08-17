"""Streaming Excel reader for million-row files."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def read_excel_rows(
    file_path: str | Path,
    sheet: str | None = None,
    header_row: int = 1,
    batch_size: int = 100,
):
    """Stream rows from an Excel file as dictionaries.

    Yields batches of rows to keep memory bounded.

    Args:
        file_path: Path to the Excel file.
        sheet: Sheet name (defaults to active sheet).
        header_row: Row number containing headers (1-indexed).
        batch_size: Number of rows per batch.

    Yields:
        List of dicts, one per row, keyed by header names.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    headers = []
    batch = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == header_row:
            headers = [str(h) if h is not None else f"col_{j}" for j, h in enumerate(row)]
            continue

        if i < header_row:
            continue

        if not any(row):
            continue

        row_dict = {}
        for j, val in enumerate(row):
            key = headers[j] if j < len(headers) else f"col_{j}"
            row_dict[key] = val
        row_dict["__row_index"] = i

        batch.append(row_dict)
        if len(batch) >= batch_size:
            yield batch
            batch = []

    if batch:
        yield batch

    wb.close()


def get_row_count(file_path: str | Path, sheet: str | None = None) -> int:
    """Get total row count from an Excel file."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    count = ws.max_row or 0
    wb.close()
    return count
