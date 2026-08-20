"""Streaming Excel writer for appending row results.

Uses openpyxl to create/update Excel files with result columns including
research data (specs, images, videos) and usage metrics (tokens, API calls,
elapsed time). Creates workbook with styled headers on first write.

Columns: row_index, product_name, status, confidence, specifications,
source_urls, image_urls, image_paths, image_views, video_urls, video_paths,
error, input_tokens, output_tokens, total_tokens, llm_calls, serpapi_calls,
elapsed_seconds, missing_views.
"""

from __future__ import annotations

import json
import logging
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

RESULT_COLUMNS = [
    "row_index",
    "product_name",
    "status",
    "confidence",
    "specifications",
    "source_urls",
    "image_urls",
    "image_paths",
    "image_views",
    "video_urls",
    "video_paths",
    "error",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    "llm_calls",
    "serpapi_calls",
    "elapsed_seconds",
    "missing_views",
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

COLUMN_WIDTHS = {
    "A": 12,   # row_index
    "B": 40,   # product_name
    "C": 12,   # status
    "D": 12,   # confidence
    "E": 50,   # specifications
    "F": 60,   # source_urls
    "G": 60,   # image_urls
    "H": 60,   # image_paths
    "I": 30,   # image_views
    "J": 60,   # video_urls
    "K": 60,   # video_paths
    "L": 40,   # error
    "M": 14,   # input_tokens
    "N": 14,   # output_tokens
    "O": 14,   # total_tokens
    "P": 12,   # llm_calls
    "Q": 14,   # serpapi_calls
    "R": 14,   # elapsed_seconds
    "S": 40,   # missing_views
}


def _ensure_workbook(file_path: str | Path) -> None:
    """Create workbook with headers if it doesn't exist."""
    if os.path.exists(file_path):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    for col, header in enumerate(RESULT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(file_path)
    wb.close()


def _serialize_value(value) -> str:
    """Serialize a value for Excel storage."""
    if isinstance(value, (list, dict)):
        return json.dumps(value, default=str)
    if value is None:
        return ""
    return str(value)


def write_row_result(file_path: str | Path, result: dict) -> None:
    """Append a single row result to the Excel file.

    Args:
        file_path: Path to the Excel file.
        result: Dict with keys matching RESULT_COLUMNS.
    """
    file_path = str(file_path)
    _ensure_workbook(file_path)

    wb = load_workbook(file_path)
    ws = wb.active

    next_row = ws.max_row + 1 if ws.max_row else 2

    for col, key in enumerate(RESULT_COLUMNS, start=1):
        value = result.get(key, "")
        ws.cell(row=next_row, column=col, value=_serialize_value(value))

    wb.save(file_path)
    wb.close()


def write_row_results(file_path: str | Path, results: list[dict]) -> None:
    """Append multiple row results to the Excel file.

    Args:
        file_path: Path to the Excel file.
        results: List of dicts with keys matching RESULT_COLUMNS.
    """
    if not results:
        return

    file_path = str(file_path)
    _ensure_workbook(file_path)

    wb = load_workbook(file_path)
    ws = wb.active

    next_row = ws.max_row + 1 if ws.max_row else 2

    for result in results:
        for col, key in enumerate(RESULT_COLUMNS, start=1):
            value = result.get(key, "")
            ws.cell(row=next_row, column=col, value=_serialize_value(value))
        next_row += 1

    wb.save(file_path)
    wb.close()
